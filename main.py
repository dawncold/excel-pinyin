import sys
import os.path
from openpyxl import load_workbook
from pypinyin import lazy_pinyin

def format_pinyin(pinyins):
    return f"{pinyins[0]} {''.join(pinyins[1:])}".strip()

if __name__ == '__main__':
    file_path = sys.argv[1]

    wb = load_workbook(file_path)
    ws = wb.active

    for i in range(ws.max_row):
        cname = ws.cell(column=2, row=i+1)
        pinyins = lazy_pinyin(cname.value)
        ws.cell(column=3, row=i+1, value=format_pinyin(pinyins))
    
    p, ext = os.path.splitext(file_path)
    wb.save(f'{p}-new{ext}')